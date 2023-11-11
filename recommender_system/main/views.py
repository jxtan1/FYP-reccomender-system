from django.shortcuts import render, redirect, get_object_or_404
# from .forms import CustomerFeedbackForm, SellerFeedbackForm, RegisterForm, ProductForm
# from .forms import RegisterForm, ProductForm, UserProfileEditForm
from .forms import *
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, logout, authenticate
from .models import CustomUser, Product, Review, Cart, CartItem#, Reviewer 
from django.core.paginator import Paginator
from django.db.models import Q
from django.db.models import Avg
from django.views.generic import ListView
from django.urls import reverse
from django.http import HttpResponseRedirect

from django.contrib.auth.views import PasswordChangeView
from django.urls import reverse_lazy
from django.views.generic.base import TemplateView
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
from django.views.generic.edit import DeleteView
from django.contrib.auth.models import User
from .forms import ReviewForm


from django.shortcuts import redirect


# Create your views here.
# In views.py
from django.contrib.auth.decorators import user_passes_test

def is_admin(user):
    return user.is_authenticated and user.is_staff  # Check if the user is authenticated and is a staff member (admin).

def is_customer_or_seller(user):
    # Check if the user is a customer or a seller
    return user.is_authenticated and user.account in ('B', 'S')


def is_customer(user):
    # Check if the user is a customer or a seller
    return user.is_authenticated and user.account in ('B')

def is_seller(user):
    # Check if the user is a customer or a seller
    return user.is_authenticated and user.account in ('S')

@login_required(login_url="/login")
def home(request):
    all_products = Product.objects.order_by('product_id')
    paginator = Paginator(all_products, 50)

    page = request.GET.get('page')
    products = paginator.get_page(page)

    # Calculate the average rating for each product
    for product in products:
        product.average_rating = Review.objects.filter(product_name=product).aggregate(avg_rating=Avg('rating'))['avg_rating']

    return render(request, 'main/home.html', {"products": products})


def sign_up(request):
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            #login(request, user)
            return render(request, 'registration/sign_up_success.html')
        else:
            for key, error in list(form.errors.items()):
                if key == 'captcha' and error[0] == 'This field is required.':
                    messages.error(request, "You must pass the reCAPTCHA test")
                    continue
                
                messages.error(request, error) 
    else:
        form = RegisterForm()

    return render(request, 'registration/sign_up.html', {"form": form})


@login_required(login_url="/login")
@user_passes_test(is_seller, login_url="/home")
def add_product(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            product = form.save(commit=False)
            # Automatically set the seller to the current user (seller)
            product.seller = request.user
            product.save()
            return render(request, 'main/add_product_success.html')
    else:
        form = ProductForm()

    return render(request, 'main/add_product.html', {"form": form})


class SellerProductListView(ListView):
    model = Product
    template_name = 'main/seller_my_products.html'
    context_object_name = 'products'

    def get_queryset(self):
        # Filter products to show only the ones owned by the logged-in seller
        return Product.objects.filter(seller=self.request.user)
    

def edit_product(request, product_id):
    product = get_object_or_404(Product, pk=product_id)

    if request.user != product.seller:
        # Check if the product belongs to the currently logged-in seller
        # If not, return a permission denied or error message
        return render(request, 'main/not_authorized.html')

    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
        return render(request, 'main/edit_product_success.html')
    else:
        form = ProductForm(instance=product)

    return render(request, 'main/edit_product.html', {'form': form, 'product': product})

def delete_product(request, product_id):
    product = get_object_or_404(Product, pk=product_id)

    if request.user != product.seller:
        # Check if the product belongs to the currently logged-in seller
        # If not, return a permission denied or error message
        return render(request, 'main/not_authorized.html')

    if request.method == 'POST':
        product.delete()
        return render(request, 'main/delete_product_success.html')

    return render(request, 'main/delete_product_confirm.html', {'product': product})






def product_detail(request, product_id):
    product = get_object_or_404(Product, pk=product_id)
    reviews = Review.objects.filter(product_name=product)

    total_ratings = reviews.count()  # Calculate the total number of ratings

    # Calculate the average rating
    if total_ratings > 0:
        average_rating = sum(review.rating for review in reviews) / total_ratings
    else:
        average_rating = 0  # Default to 0 if there are no ratings

    # Create a Paginator object for reviews
    paginator = Paginator(reviews, 10)  # Change 10 to the number of reviews per page you desire

    # Get the current page number from the request's GET parameters
    page = request.GET.get('page')

    # Get the Page object for the current page of reviews
    reviews = paginator.get_page(page)

    # Review form handling
    if request.method == 'POST':
        form = ReviewForm(request.POST)
        if form.is_valid():
            review = form.save(commit=False)
            review.product_name = product
            review.username = request.user  # Assuming you are using Django authentication
            review.save()
            # Redirect to the success page
            return redirect('review_success', product_id=product_id)
    else:
        form = ReviewForm()

    context = {
        'product': product,
        'reviews': reviews,
        'average_rating': average_rating,
        'total_ratings': total_ratings,
        'seller': product.seller,
        'form': form,  # Add the form to the context
    }
    return render(request, 'main/product_detail.html', context)


def search_products(request):
    query = request.GET.get('q')
    
    # Filter products based on the search query
    products = Product.objects.filter(Q(name__icontains=query)).order_by('name')
    
    # Create a Paginator object
    paginator = Paginator(products, 50)  # Change 50 to the number of products per page you desire

    # Get the current page number from the request's GET parameters
    page = request.GET.get('page')

    # Get the Page object for the current page
    products = paginator.get_page(page)

    return render(request, 'main/home.html', {"products": products})



@login_required
def view_account(request):
    user = request.user  # Retrieve the current user
    return render(request, 'main/view_account.html', {'user': user})

@login_required
def edit_account(request):
    if request.method == 'POST':
        form = UserProfileEditForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            return redirect('/view_account')
    else:
        form = UserProfileEditForm(instance=request.user)
    return render(request, 'main/edit_account.html', {'form': form})


@login_required
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            # Update the session to prevent the user from being logged out
            update_session_auth_hash(request, user)
            messages.success(request, 'Your password was successfully updated!')
            return redirect('/view_account')
        else:
            messages.error(request, 'Please correct the error(s) below.')
    else:
        form = PasswordChangeForm(user=request.user)
    return render(request, 'registration/change_password.html', {'form': form})

class CustomPasswordChangeView(PasswordChangeView):
    success_url = reverse_lazy('password_change_done')  # Redirect to a success page after changing the password

class PasswordChangeDoneView(TemplateView):
    template_name = 'registration/password_change_done.html'

    @login_required
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)


@login_required
def delete_account(request):
    if request.method == 'POST' and 'confirm_delete' in request.POST:
        # Delete the user account
        user = request.user
        user.delete()

        return redirect('logout')  # You can redirect to the logout view or any other appropriate page
    return render(request, 'main/confirm_delete.html')



# from openpyxl import load_workbook
import pandas as pd
from django.db import IntegrityError

@login_required(login_url="/login")
@user_passes_test(is_admin, login_url="/home")
def import_from_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']

        # Read file and convert to pandas dataframe
        df = pd.read_excel(excel_file)
        df.drop_duplicates() # Cleaning
        
        # Handle Product data
        if request.POST.get('import_type') == "Product":
            #df.pop(df.columns[-1]) # Remove urls from dataset
            df.pop('Product Page')
            # Rename columns to match model
            rename_columns = {"Product Name": "name", "Price": "price", "Total Sold": "sold_count", "Seller": 'seller', "Description": "description", "Image": "image"}
            df.rename(columns = rename_columns, inplace=True)
            df['price'] = df['price'].astype(str).str.replace(',', '').astype(float) # Change sold_count to int
            df['sold_count'] = df['sold_count'].astype(str).str.replace(',', '').astype(int) # Change sold_count to int
            sellers = (df.loc[:,'seller']).drop_duplicates() # Get unique sellers
            # Save all sellers from products gathered to the Seller model
            for item in sellers:
                # Need a way to catch verify unique=True error
                try:
                    seller = CustomUser(username=item, account='S') # Should not update any existing due to unique constraint
                    seller.save()
                except IntegrityError as e:
                    if 'unique constraint' in e.args[0]:
                        #Skip repeats
                        print(e)
            data = df.to_dict('records') # Convert dataframe into dictionaries
            # Loop over list of dictionaries and save each one to the Product model
            for item in data:
                try:
                    seller = CustomUser.objects.get(username=item['seller'], account='S')
                except CustomUser.DoesNotExist:
                    # Skip sellers not in the database
                    print(item[''], "not found!")
                    continue
                try:
                    product = Product.objects.update_or_create(name=item['name'], price=item['price'], sold_count=item['sold_count'], seller=seller, description=item['description'], image=item['image'])
                    product[0].save()
                except IntegrityError as e:
                    if 'unique constraint' in e.args[0]:
                        # Skip repeats
                        print('Duplicate', item['name'], 'found, ignoring...')
        
        # Handle Review data
        elif request.POST.get('import_type') == "Review":
            # Rename columns to match review model
            rename_columns = {"Product Name": "product_name", "Rating": "rating", "Reviewer": "username", "Content": "comment"}
            df.rename(columns = rename_columns, inplace=True)
            reviewers = (df.loc[:,'username']).drop_duplicates() # Get unique reviewers
            # Save all reviewers from the reviews gathered to the seller (CustomUser account='B') model
            for item in reviewers:
                # Need a way to catch verify unique=True error
                try:
                    reviewer = CustomUser(username=item, account='B') # Should not update any existing due to unique constraint
                    reviewer.save()
                except IntegrityError as e:
                    if 'unique constraint' in e.args[0]:
                        # Skip repeats
                        print(e)
            data = df.to_dict('records') # Convert dataframe into dictionaries
            # Loop over list of dictionaries and save each one to the Product model
            for item in data:
                # Find product & seller (CustomUser Account='B') objects
                try:
                    product = Product.objects.get(name=item["product_name"])
                except Product.DoesNotExist:
                    # Skip products not in database
                    print(item["product_name"], "not found!")
                    continue
                try:
                    reviewer = CustomUser.objects.get(username=item["username"], account='B')
                except CustomUser.DoesNotExist:
                    # Skip if cannot find reviewer
                    continue
                review = Review.objects.update_or_create(product_name=product, rating=item["rating"], username=reviewer, comment=item["comment"])
                review[0].save()
        
        return render(request, 'main/import_success.html')
    return render(request, 'main/import_form.html')

@login_required(login_url="/login")
@user_passes_test(is_admin, login_url="/home")
def admin_force_gen(request):
    return render(request, 'main/admin_force_gen.html')

from django.db.models import Count

@login_required(login_url="/login")
@user_passes_test(is_customer, login_url="/home")
def buyer_recommendations(request):
    #gen_similarity_matrix() # Only if you want to regenerate matrix everytime recommendations are viewed
    prodlist = []
    try:
        results = make_prediction() # list of [username, product, prediction]
        for i in range(len(results)):
            print("Predicted", results[i][0], "rating", results[i][1], "as", results[i][2])
            #printout.append("Predicted {0} rating {1} as {2}".format(results[i][0], results[i][1], results[i][2]))
            prodlist.append(results[i][1])
        #return printout
    except:
        print("Not enough data to recommend user, loading top 10 items")
        # Find products with the top 10 highest ratings
        top10 = Review.objects.values('product_name').annotate(Count('rating')).order_by('-rating__count')[:10]
        for i in range(top10.count()):
            print(top10)
            prodlist.append(top10[i]['product_name'])
    
    products = Product.objects.filter(name__in=prodlist)
    # Calculate the average rating for each product
    for product in products:
        product.average_rating = Review.objects.filter(product_name=product).aggregate(avg_rating=Avg('rating'))['avg_rating']
    context = {
        'seller': product.seller,
        'products': products,
        'predicted': ""
    }
    return render(request, 'main/seller_store.html', context)

from collaborative_filtering.recommender import *

@login_required(login_url="/login")
@user_passes_test(is_seller, login_url="/home")
def seller_predict_buyer(request):
    if request.method == 'POST':
        results = []
        prodlist = []
        username = request.POST.get('username')
        similarity_thresh = request.POST.get('similarity_thresh')
        n = int(request.POST.get('n'))
        m = int(request.POST.get('m'))

        # test_accuracy(n, m)
    
        try:
            results = make_prediction(username, similarity_thresh, n, m) # list of [username, product, prediction]
            for i in range(len(results)):
                print("Predicted", results[i][0], "rating", results[i][1], "as", results[i][2])
                #printout.append("Predicted {0} rating {1} as {2}".format(results[i][0], results[i][1], results[i][2]))
                prodlist.append(results[i][1])
            #return printout
        except:
            print("Not enough data to recommend user, loading top {m} items")
            # Find products with the top m highest ratings
            topm = Review.objects.values('product_name').annotate(Count('rating')).order_by('-rating__count')[:m]
            for i in range(topm.count()):
                print(topm)
                prodlist.append(topm[i]['product_name'])
        
        products = Product.objects.filter(name__in=prodlist)
        # Calculate the average rating for each product
        for i in range(len(products)):
            product = products[i]
            product.average_rating = Review.objects.filter(product_name=product).aggregate(avg_rating=Avg('rating'))['avg_rating']
            try:
                product.predicted_rating = results[i][2]
            except:
                product.predicted_rating = Review.objects.filter(product_name=product).aggregate(avg_rating=Avg('rating'))['avg_rating']
        predict = "Predicted Rating: "
        context = {
            'seller': product.seller,
            'products': products,
            'predict': predict,
        }
        return render(request, 'main/seller_store.html', context)
    
    return render(request, 'main/seller_predict_buyer.html')

@login_required(login_url="/login")
@user_passes_test(is_customer, login_url="/home")
def customer_feedback_view(request):
    if request.method == 'POST':
        form = CustomerFeedbackForm(request.POST)
        if form.is_valid():
            feedback = form.save(commit=False)
            feedback.respondent = request.user  # Automatically set the user
            feedback.save()
            # Redirect or display a thank you message
            return render(request, 'main/feedback_thankyou_message.html')

    else:
        form = CustomerFeedbackForm()

    return render(request, 'main/feedback_form.html', {'form': form})


@login_required(login_url="/login")
@user_passes_test(is_seller, login_url="/home")
def seller_feedback_view(request):
    if request.method == 'POST':
        form = SellerFeedbackForm(request.POST)
        if form.is_valid():
            feedback = form.save(commit=False)
            feedback.respondent = request.user  # Automatically set the user
            feedback.save()
            # Redirect or display a thank you message
            return render(request, 'main/feedback_thankyou_message.html')

    else:
        form = SellerFeedbackForm()

    return render(request, 'main/feedback_form.html', {'form': form})


def seller_store(request, username):
    seller = get_object_or_404(CustomUser, username=username)
    products = Product.objects.filter(seller=seller)
    # Calculate the average rating for each product
    for product in products:
        product.average_rating = Review.objects.filter(product_name=product).aggregate(avg_rating=Avg('rating'))['avg_rating']
    context = {
        'seller': seller,
        'products': products,
    }
    return render(request, 'main/seller_store.html', context)


# Common function to update the cart and retrieve cart items
def update_cart(request):
    cart = None
    cartitems = []

    if request.user.is_authenticated:
        cart, created = Cart.objects.get_or_create(user=request.user, completed=False)
        cartitems = cart.cartitems.all()

    context = {"cart": cart, "items": cartitems}
    return context

@login_required(login_url="/login")
@user_passes_test(is_customer, login_url="/home")
def cart(request):
    context = update_cart(request)
    return render(request, 'main/cart.html', context)


@login_required(login_url="/login")
@user_passes_test(is_customer, login_url="/home")
def add_to_cart(request, product_id):
    product = get_object_or_404(Product, product_id=product_id)

    if request.user.is_authenticated:
        cart, created = Cart.objects.get_or_create(user=request.user, completed=False)
        cartitem, created = CartItem.objects.get_or_create(cart=cart, product=product)
        cartitem.quantity += 1
        cartitem.save()
        #messages.success(request, 'You have added an item to your cart successfully')

    # Call the common function to update the cart
    context = update_cart(request)

    return render(request, 'main/cart.html', context)


def remove_from_cart(request, product_id):
    product = get_object_or_404(Product, pk=product_id)
    cart, created = Cart.objects.get_or_create(user=request.user)
    cartitem = CartItem.objects.filter(cart=cart, product=product).first()

    if cartitem:
        cartitem.quantity -= 1

        if cartitem.quantity <= 0:
            cartitem.delete()
        else:
            cartitem.save()

    # Call the common function to update the cart
    context = update_cart(request)

    return render(request, 'main/cart.html', context)

def clear_cart(request):
    cart, created = Cart.objects.get_or_create(user=request.user)
    cart.delete()
    # Call the common function to update the cart
    context = update_cart(request)

    return render(request, 'main/cart.html', context)

def checkout(request):
    if request.method == 'POST':
        payment_form = PaymentForm(request.POST)

        if payment_form.is_valid():
            cart, created = Cart.objects.get_or_create(user=request.user)
            cart.delete() # Clear the cart
            return render(request, 'main/payment_success.html')

    else:
        payment_form = PaymentForm()

    context = {'payment_form': payment_form}
    return render(request, 'main/payment.html', context)

def review_success(request, product_id):
    product = get_object_or_404(Product, pk=product_id)
    return render(request, 'main/review_success.html', {'product': product})



from .forms import CustomerFeedbackForm, SellerFeedbackForm, RegisterForm, ProductForm
from .forms import RegisterForm, ProductForm, UserProfileEditForm, PaymentForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, logout, authenticate
from .models import CustomUser, Order, OrderItem, Payment, Product, Review, Cart, CartItem#, Reviewer 
from django.core.paginator import Paginator
from django.db.models import Q
from django.db.models import Avg
from django.views.generic import ListView
from django.urls import reverse
from django.http import HttpResponseRedirect

from django.contrib.auth.views import PasswordChangeView
from django.urls import reverse_lazy
from django.views.generic.base import TemplateView
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
from django.views.generic.edit import DeleteView
from django.contrib.auth.models import User
from .forms import ReviewForm

# Create your views here.
# In views.py
from django.contrib.auth.decorators import user_passes_test

def is_admin(user):
    return user.is_authenticated and user.is_staff  # Check if the user is authenticated and is a staff member (admin).

def is_customer_or_seller(user):
    # Check if the user is a customer or a seller
    return user.is_authenticated and user.account in ('B', 'S')


def is_customer(user):
    # Check if the user is a customer or a seller
    return user.is_authenticated and user.account in ('B')

def is_seller(user):
    # Check if the user is a customer or a seller
    return user.is_authenticated and user.account in ('S')

@login_required(login_url="/login")
def home(request):
    all_products = Product.objects.order_by('product_id')
    paginator = Paginator(all_products, 50)

    page = request.GET.get('page')
    products = paginator.get_page(page)

    # Calculate the average rating for each product
    for product in products:
        product.average_rating = Review.objects.filter(product_name=product).aggregate(avg_rating=Avg('rating'))['avg_rating']

    return render(request, 'main/home.html', {"products": products})


def sign_up(request):
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            #login(request, user)
            return render(request, 'registration/sign_up_success.html')
        else:
            for key, error in list(form.errors.items()):
                if key == 'captcha' and error[0] == 'This field is required.':
                    messages.error(request, "You must pass the reCAPTCHA test")
                    continue
                
                messages.error(request, error) 
    else:
        form = RegisterForm()

    return render(request, 'registration/sign_up.html', {"form": form})


@login_required(login_url="/login")
@user_passes_test(is_seller, login_url="/home")
def add_product(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            product = form.save(commit=False)
            # Automatically set the seller to the current user (seller)
            product.seller = request.user
            product.save()
            return render(request, 'main/add_product_success.html')
    else:
        form = ProductForm()

    return render(request, 'main/add_product.html', {"form": form})


class SellerProductListView(ListView):
    model = Product
    template_name = 'main/seller_my_products.html'
    context_object_name = 'products'

    def get_queryset(self):
        # Filter products to show only the ones owned by the logged-in seller
        return Product.objects.filter(seller=self.request.user)
    

def edit_product(request, product_id):
    product = get_object_or_404(Product, pk=product_id)

    if request.user != product.seller:
        # Check if the product belongs to the currently logged-in seller
        # If not, return a permission denied or error message
        return render(request, 'main/not_authorized.html')

    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
        return render(request, 'main/edit_product_success.html')
    else:
        form = ProductForm(instance=product)

    return render(request, 'main/edit_product.html', {'form': form, 'product': product})

def delete_product(request, product_id):
    product = get_object_or_404(Product, pk=product_id)

    if request.user != product.seller:
        # Check if the product belongs to the currently logged-in seller
        # If not, return a permission denied or error message
        return render(request, 'main/not_authorized.html')

    if request.method == 'POST':
        product.delete()
        return render(request, 'main/delete_product_success.html')

    return render(request, 'main/delete_product_confirm.html', {'product': product})






def product_detail(request, product_id):
    product = get_object_or_404(Product, pk=product_id)
    reviews = Review.objects.filter(product_name=product)

    total_ratings = reviews.count()  # Calculate the total number of ratings

    # Calculate the average rating
    if total_ratings > 0:
        average_rating = sum(review.rating for review in reviews) / total_ratings
    else:
        average_rating = 0  # Default to 0 if there are no ratings

    # Create a Paginator object for reviews
    paginator = Paginator(reviews, 10)  # Change 10 to the number of reviews per page you desire

    # Get the current page number from the request's GET parameters
    page = request.GET.get('page')

    # Get the Page object for the current page of reviews
    reviews = paginator.get_page(page)

    # Review form handling
    if request.method == 'POST':
        form = ReviewForm(request.POST)
        if form.is_valid():
            review = form.save(commit=False)
            review.product_name = product
            review.username = request.user  # Assuming you are using Django authentication
            review.save()
    else:
        form = ReviewForm()

    context = {
        'product': product,
        'reviews': reviews,
        'average_rating': average_rating,
        'total_ratings': total_ratings,
        'seller': product.seller,
        'form': form,  # Add the form to the context
    }
    return render(request, 'main/product_detail.html', context)


def search_products(request):
    query = request.GET.get('q')
    
    # Filter products based on the search query
    products = Product.objects.filter(Q(name__icontains=query)).order_by('name')
    
    # Create a Paginator object
    paginator = Paginator(products, 50)  # Change 50 to the number of products per page you desire

    # Get the current page number from the request's GET parameters
    page = request.GET.get('page')

    # Get the Page object for the current page
    products = paginator.get_page(page)

    return render(request, 'main/home.html', {"products": products})



@login_required
def view_account(request):
    user = request.user  # Retrieve the current user
    return render(request, 'main/view_account.html', {'user': user})

@login_required
def edit_account(request):
    if request.method == 'POST':
        form = UserProfileEditForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            return redirect('/view_account')
    else:
        form = UserProfileEditForm(instance=request.user)
    return render(request, 'main/edit_account.html', {'form': form})


@login_required
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            # Update the session to prevent the user from being logged out
            update_session_auth_hash(request, user)
            messages.success(request, 'Your password was successfully updated!')
            return redirect('/view_account')
        else:
            messages.error(request, 'Please correct the error(s) below.')
    else:
        form = PasswordChangeForm(user=request.user)
    return render(request, 'registration/change_password.html', {'form': form})

class CustomPasswordChangeView(PasswordChangeView):
    success_url = reverse_lazy('password_change_done')  # Redirect to a success page after changing the password

class PasswordChangeDoneView(TemplateView):
    template_name = 'registration/password_change_done.html'

    @login_required
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)


@login_required
def delete_account(request):
    if request.method == 'POST' and 'confirm_delete' in request.POST:
        # Delete the user account
        user = request.user
        user.delete()

        return redirect('logout')  # You can redirect to the logout view or any other appropriate page
    return render(request, 'main/confirm_delete.html')



# from openpyxl import load_workbook
import pandas as pd
from django.db import IntegrityError

@login_required(login_url="/login")
@user_passes_test(is_admin, login_url="/home")
def import_from_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']

        # Read file and convert to pandas dataframe
        df = pd.read_excel(excel_file)
        df.drop_duplicates() # Cleaning
        
        # Handle Product data
        if request.POST.get('import_type') == "Product":
            #df.pop(df.columns[-1]) # Remove urls from dataset
            df.pop('Product Page')
            # Rename columns to match model
            rename_columns = {"Product Name": "name", "Price": "price", "Total Sold": "sold_count", "Seller": 'seller', "Description": "description", "Image": "image"}
            df.rename(columns = rename_columns, inplace=True)
            df['price'] = df['price'].astype(str).str.replace(',', '').astype(float) # Change sold_count to int
            df['sold_count'] = df['sold_count'].astype(str).str.replace(',', '').astype(int) # Change sold_count to int
            sellers = (df.loc[:,'seller']).drop_duplicates() # Get unique sellers
            # Save all sellers from products gathered to the Seller model
            for item in sellers:
                # Need a way to catch verify unique=True error
                try:
                    seller = CustomUser(username=item, account='S') # Should not update any existing due to unique constraint
                    seller.save()
                except IntegrityError as e:
                    if 'unique constraint' in e.args[0]:
                        #Skip repeats
                        print(e)
            data = df.to_dict('records') # Convert dataframe into dictionaries
            # Loop over list of dictionaries and save each one to the Product model
            for item in data:
                try:
                    seller = CustomUser.objects.get(username=item['seller'], account='S')
                except CustomUser.DoesNotExist:
                    # Skip sellers not in the database
                    print(item[''], "not found!")
                    continue
                try:
                    product = Product.objects.update_or_create(name=item['name'], price=item['price'], sold_count=item['sold_count'], seller=seller, description=item['description'], image=item['image'])
                    product[0].save()
                except IntegrityError as e:
                    if 'unique constraint' in e.args[0]:
                        # Skip repeats
                        print('Duplicate', item['name'], 'found, ignoring...')
        
        # Handle Review data
        elif request.POST.get('import_type') == "Review":
            # Rename columns to match review model
            rename_columns = {"Product Name": "product_name", "Rating": "rating", "Reviewer": "username", "Content": "comment"}
            df.rename(columns = rename_columns, inplace=True)
            reviewers = (df.loc[:,'username']).drop_duplicates() # Get unique reviewers
            # Save all reviewers from the reviews gathered to the seller (CustomUser account='B') model
            for item in reviewers:
                # Need a way to catch verify unique=True error
                try:
                    reviewer = CustomUser(username=item, account='B') # Should not update any existing due to unique constraint
                    reviewer.save()
                except IntegrityError as e:
                    if 'unique constraint' in e.args[0]:
                        # Skip repeats
                        print(e)
            data = df.to_dict('records') # Convert dataframe into dictionaries
            # Loop over list of dictionaries and save each one to the Product model
            for item in data:
                # Find product & seller (CustomUser Account='B') objects
                try:
                    product = Product.objects.get(name=item["product_name"])
                except Product.DoesNotExist:
                    # Skip products not in database
                    print(item["product_name"], "not found!")
                    continue
                try:
                    reviewer = CustomUser.objects.get(username=item["username"], account='B')
                except CustomUser.DoesNotExist:
                    # Skip if cannot find reviewer
                    continue
                review = Review.objects.update_or_create(product_name=product, rating=item["rating"], username=reviewer, comment=item["comment"])
                review[0].save()
        
        return render(request, 'main/import_success.html')
    return render(request, 'main/import_form.html')

@login_required(login_url="/login")
@user_passes_test(is_admin, login_url="/home")
def admin_force_gen(request):
    return render(request, 'main/admin_force_gen.html')

from django.db.models import Count

@login_required(login_url="/login")
@user_passes_test(is_customer, login_url="/home")
def buyer_recommendations(request):
    #gen_similarity_matrix() # Only if you want to regenerate matrix everytime recommendations are viewed
    prodlist = []
    try:
        results = make_prediction() # list of [username, product, prediction]
        for i in range(len(results)):
            print("Predicted", results[i][0], "rating", results[i][1], "as", results[i][2])
            #printout.append("Predicted {0} rating {1} as {2}".format(results[i][0], results[i][1], results[i][2]))
            prodlist.append(results[i][1])
        #return printout
    except:
        print("Not enough data to recommend user, loading top 10 items")
        # Find products with the top 10 highest ratings
        top10 = Review.objects.values('product_name').annotate(Count('rating')).order_by('-rating__count')[:10]
        for i in range(top10.count()):
            print(top10)
            prodlist.append(top10[i]['product_name'])
    
    products = Product.objects.filter(name__in=prodlist)
    # Calculate the average rating for each product
    for product in products:
        product.average_rating = Review.objects.filter(product_name=product).aggregate(avg_rating=Avg('rating'))['avg_rating']
    context = {
        'seller': product.seller,
        'products': products,
        'predicted': ""
    }
    return render(request, 'main/seller_store.html', context)

from collaborative_filtering.recommender import *

@login_required(login_url="/login")
@user_passes_test(is_seller, login_url="/home")
def seller_predict_buyer(request):
    if request.method == 'POST':
        results = []
        prodlist = []
        username = request.POST.get('username')
        similarity_thresh = request.POST.get('similarity_thresh')
        n = int(request.POST.get('n'))
        m = int(request.POST.get('m'))

        # test_accuracy(n, m)
    
        try:
            results = make_prediction(username, similarity_thresh, n, m) # list of [username, product, prediction]
            for i in range(len(results)):
                print("Predicted", results[i][0], "rating", results[i][1], "as", results[i][2])
                #printout.append("Predicted {0} rating {1} as {2}".format(results[i][0], results[i][1], results[i][2]))
                prodlist.append(results[i][1])
            #return printout
        except:
            print("Not enough data to recommend user, loading top {m} items")
            # Find products with the top m highest ratings
            topm = Review.objects.values('product_name').annotate(Count('rating')).order_by('-rating__count')[:m]
            for i in range(topm.count()):
                print(topm)
                prodlist.append(topm[i]['product_name'])
        
        products = Product.objects.filter(name__in=prodlist)
        # Calculate the average rating for each product
        for i in range(len(products)):
            product = products[i]
            product.average_rating = Review.objects.filter(product_name=product).aggregate(avg_rating=Avg('rating'))['avg_rating']
            try:
                product.predicted_rating = results[i][2]
            except:
                product.predicted_rating = Review.objects.filter(product_name=product).aggregate(avg_rating=Avg('rating'))['avg_rating']
        predict = "Predicted Rating: "
        context = {
            'seller': product.seller,
            'products': products,
            'predict': predict,
        }
        return render(request, 'main/seller_store.html', context)
    
    return render(request, 'main/seller_predict_buyer.html')

@login_required(login_url="/login")
@user_passes_test(is_customer, login_url="/home")
def customer_feedback_view(request):
    if request.method == 'POST':
        form = CustomerFeedbackForm(request.POST)
        if form.is_valid():
            feedback = form.save(commit=False)
            feedback.respondent = request.user  # Automatically set the user
            feedback.save()
            # Redirect or display a thank you message
            return render(request, 'main/feedback_thankyou_message.html')

    else:
        form = CustomerFeedbackForm()

    return render(request, 'main/feedback_form.html', {'form': form})


@login_required(login_url="/login")
@user_passes_test(is_seller, login_url="/home")
def seller_feedback_view(request):
    if request.method == 'POST':
        form = SellerFeedbackForm(request.POST)
        if form.is_valid():
            feedback = form.save(commit=False)
            feedback.respondent = request.user  # Automatically set the user
            feedback.save()
            # Redirect or display a thank you message
            return render(request, 'main/feedback_thankyou_message.html')

    else:
        form = SellerFeedbackForm()

    return render(request, 'main/feedback_form.html', {'form': form})


def seller_store(request, username):
    seller = get_object_or_404(CustomUser, username=username)
    products = Product.objects.filter(seller=seller)
    # Calculate the average rating for each product
    for product in products:
        product.average_rating = Review.objects.filter(product_name=product).aggregate(avg_rating=Avg('rating'))['avg_rating']
    context = {
        'seller': seller,
        'products': products,
    }
    return render(request, 'main/seller_store.html', context)


# Common function to update the cart and retrieve cart items
def update_cart(request):
    cart = None
    cartitems = []

    if request.user.is_authenticated:
        cart, created = Cart.objects.get_or_create(user=request.user, completed=False)
        cartitems = cart.cartitems.all()

    context = {"cart": cart, "items": cartitems}
    return context

@login_required(login_url="/login")
@user_passes_test(is_customer, login_url="/home")
def cart(request):
    context = update_cart(request)
    return render(request, 'main/cart.html', context)


@login_required(login_url="/login")
@user_passes_test(is_customer, login_url="/home")
def add_to_cart(request, product_id):
    product = get_object_or_404(Product, product_id=product_id)

    if request.user.is_authenticated:
        cart, created = Cart.objects.get_or_create(user=request.user, completed=False)
        cartitem, created = CartItem.objects.get_or_create(cart=cart, product=product)
        cartitem.quantity += 1
        cartitem.save()
        #messages.success(request, 'You have added an item to your cart successfully')

    # Call the common function to update the cart
    context = update_cart(request)

    return render(request, 'main/cart.html', context)

def remove_from_cart(request, product_id):
    product = get_object_or_404(Product, pk=product_id)
    cart, created = Cart.objects.get_or_create(user=request.user)
    cartitem = CartItem.objects.filter(cart=cart, product=product).first()

    if cartitem:
        cartitem.quantity -= 1

        if cartitem.quantity <= 0:
            cartitem.delete()
        else:
            cartitem.save()

    # Call the common function to update the cart
    context = update_cart(request)

    return render(request, 'main/cart.html', context)

def clear_cart(request):
    cart, created = Cart.objects.get_or_create(user=request.user)
    cart.delete()
    # Call the common function to update the cart
    context = update_cart(request)

    return render(request, 'main/cart.html', context)

def checkout(request):
    if request.method == 'POST':
        payment_form = PaymentForm(request.POST)

        if payment_form.is_valid():
            cart, created = Cart.objects.get_or_create(user=request.user)
            cart.delete() # Clear the cart
            return render(request, 'main/payment_success.html')

    else:
        payment_form = PaymentForm()

    context = {'payment_form': payment_form}
    return render(request, 'main/payment.html', context)

def review_success(request, product_id):
    product = get_object_or_404(Product, pk=product_id)
    return render(request, 'main/review_success.html', {'product': product})



from .forms import CustomerFeedbackForm, SellerFeedbackForm, RegisterForm, ProductForm
from .forms import RegisterForm, ProductForm, UserProfileEditForm, PaymentForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, logout, authenticate
from .models import CustomUser, Order, OrderItem, Payment, Product, Review, Cart, CartItem#, Reviewer 
from django.core.paginator import Paginator
from django.db.models import Q
from django.db.models import Avg
from django.views.generic import ListView
from django.urls import reverse
from django.http import HttpResponseRedirect

from django.contrib.auth.views import PasswordChangeView
from django.urls import reverse_lazy
from django.views.generic.base import TemplateView
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
from django.views.generic.edit import DeleteView
from django.contrib.auth.models import User
from .forms import ReviewForm

# Create your views here.
# In views.py
from django.contrib.auth.decorators import user_passes_test

def is_admin(user):
    return user.is_authenticated and user.is_staff  # Check if the user is authenticated and is a staff member (admin).

def is_customer_or_seller(user):
    # Check if the user is a customer or a seller
    return user.is_authenticated and user.account in ('B', 'S')


def is_customer(user):
    # Check if the user is a customer or a seller
    return user.is_authenticated and user.account in ('B')

def is_seller(user):
    # Check if the user is a customer or a seller
    return user.is_authenticated and user.account in ('S')

@login_required(login_url="/login")
def home(request):
    all_products = Product.objects.order_by('product_id')
    paginator = Paginator(all_products, 50)

    page = request.GET.get('page')
    products = paginator.get_page(page)

    # Calculate the average rating for each product
    for product in products:
        product.average_rating = Review.objects.filter(product_name=product).aggregate(avg_rating=Avg('rating'))['avg_rating']

    return render(request, 'main/home.html', {"products": products})


def sign_up(request):
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            #login(request, user)
            return render(request, 'registration/sign_up_success.html')
        else:
            for key, error in list(form.errors.items()):
                if key == 'captcha' and error[0] == 'This field is required.':
                    messages.error(request, "You must pass the reCAPTCHA test")
                    continue
                
                messages.error(request, error) 
    else:
        form = RegisterForm()

    return render(request, 'registration/sign_up.html', {"form": form})


@login_required(login_url="/login")
@user_passes_test(is_seller, login_url="/home")
def add_product(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            product = form.save(commit=False)
            # Automatically set the seller to the current user (seller)
            product.seller = request.user
            product.save()
            return render(request, 'main/add_product_success.html')
    else:
        form = ProductForm()

    return render(request, 'main/add_product.html', {"form": form})


class SellerProductListView(ListView):
    model = Product
    template_name = 'main/seller_my_products.html'
    context_object_name = 'products'

    def get_queryset(self):
        # Filter products to show only the ones owned by the logged-in seller
        return Product.objects.filter(seller=self.request.user)
    

def edit_product(request, product_id):
    product = get_object_or_404(Product, pk=product_id)

    if request.user != product.seller:
        # Check if the product belongs to the currently logged-in seller
        # If not, return a permission denied or error message
        return render(request, 'main/not_authorized.html')

    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
        return render(request, 'main/edit_product_success.html')
    else:
        form = ProductForm(instance=product)

    return render(request, 'main/edit_product.html', {'form': form, 'product': product})

def delete_product(request, product_id):
    product = get_object_or_404(Product, pk=product_id)

    if request.user != product.seller:
        # Check if the product belongs to the currently logged-in seller
        # If not, return a permission denied or error message
        return render(request, 'main/not_authorized.html')

    if request.method == 'POST':
        product.delete()
        return render(request, 'main/delete_product_success.html')

    return render(request, 'main/delete_product_confirm.html', {'product': product})






def product_detail(request, product_id):
    product = get_object_or_404(Product, pk=product_id)
    reviews = Review.objects.filter(product_name=product)

    total_ratings = reviews.count()  # Calculate the total number of ratings

    # Calculate the average rating
    if total_ratings > 0:
        average_rating = sum(review.rating for review in reviews) / total_ratings
    else:
        average_rating = 0  # Default to 0 if there are no ratings

    # Create a Paginator object for reviews
    paginator = Paginator(reviews, 10)  # Change 10 to the number of reviews per page you desire

    # Get the current page number from the request's GET parameters
    page = request.GET.get('page')

    # Get the Page object for the current page of reviews
    reviews = paginator.get_page(page)

    # Review form handling
    if request.method == 'POST':
        form = ReviewForm(request.POST)
        if form.is_valid():
            review = form.save(commit=False)
            review.product_name = product
            review.username = request.user  # Assuming you are using Django authentication
            review.save()
    else:
        form = ReviewForm()

    context = {
        'product': product,
        'reviews': reviews,
        'average_rating': average_rating,
        'total_ratings': total_ratings,
        'seller': product.seller,
        'form': form,  # Add the form to the context
    }
    return render(request, 'main/product_detail.html', context)


def search_products(request):
    query = request.GET.get('q')
    
    # Filter products based on the search query
    products = Product.objects.filter(Q(name__icontains=query)).order_by('name')
    
    # Create a Paginator object
    paginator = Paginator(products, 50)  # Change 50 to the number of products per page you desire

    # Get the current page number from the request's GET parameters
    page = request.GET.get('page')

    # Get the Page object for the current page
    products = paginator.get_page(page)

    return render(request, 'main/home.html', {"products": products})



@login_required
def view_account(request):
    user = request.user  # Retrieve the current user
    return render(request, 'main/view_account.html', {'user': user})

@login_required
def edit_account(request):
    if request.method == 'POST':
        form = UserProfileEditForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            return redirect('/view_account')
    else:
        form = UserProfileEditForm(instance=request.user)
    return render(request, 'main/edit_account.html', {'form': form})


@login_required
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            # Update the session to prevent the user from being logged out
            update_session_auth_hash(request, user)
            messages.success(request, 'Your password was successfully updated!')
            return redirect('/view_account')
        else:
            messages.error(request, 'Please correct the error(s) below.')
    else:
        form = PasswordChangeForm(user=request.user)
    return render(request, 'registration/change_password.html', {'form': form})

class CustomPasswordChangeView(PasswordChangeView):
    success_url = reverse_lazy('password_change_done')  # Redirect to a success page after changing the password

class PasswordChangeDoneView(TemplateView):
    template_name = 'registration/password_change_done.html'

    @login_required
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)


@login_required
def delete_account(request):
    if request.method == 'POST' and 'confirm_delete' in request.POST:
        # Delete the user account
        user = request.user
        user.delete()

        return redirect('logout')  # You can redirect to the logout view or any other appropriate page
    return render(request, 'main/confirm_delete.html')



from openpyxl import load_workbook
import pandas as pd
from django.db import IntegrityError

@login_required(login_url="/login")
@user_passes_test(is_admin, login_url="/home")
def import_from_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']

        # Read file and convert to pandas dataframe
        df = pd.read_excel(excel_file)
        df.drop_duplicates() # Cleaning
        
        # Handle Product data
        if request.POST.get('import_type') == "Product":
            #df.pop(df.columns[-1]) # Remove urls from dataset
            df.pop('Product Page')
            # Rename columns to match model
            rename_columns = {"Product Name": "name", "Price": "price", "Total Sold": "sold_count", "Seller": 'seller', "Description": "description", "Image": "image"}
            df.rename(columns = rename_columns, inplace=True)
            df['price'] = df['price'].astype(str).str.replace(',', '').astype(float) # Change sold_count to int
            df['sold_count'] = df['sold_count'].astype(str).str.replace(',', '').astype(int) # Change sold_count to int
            sellers = (df.loc[:,'seller']).drop_duplicates() # Get unique sellers
            # Save all sellers from products gathered to the Seller model
            for item in sellers:
                # Need a way to catch verify unique=True error
                try:
                    seller = CustomUser(username=item, account='S') # Should not update any existing due to unique constraint
                    seller.save()
                except IntegrityError as e:
                    if 'unique constraint' in e.args[0]:
                        #Skip repeats
                        print(e)
            data = df.to_dict('records') # Convert dataframe into dictionaries
            # Loop over list of dictionaries and save each one to the Product model
            for item in data:
                try:
                    seller = CustomUser.objects.get(username=item['seller'], account='S')
                except CustomUser.DoesNotExist:
                    # Skip sellers not in the database
                    print(item[''], "not found!")
                    continue
                try:
                    product = Product.objects.update_or_create(name=item['name'], price=item['price'], sold_count=item['sold_count'], seller=seller, description=item['description'], image=item['image'])
                    product[0].save()
                except IntegrityError as e:
                    if 'unique constraint' in e.args[0]:
                        # Skip repeats
                        print('Duplicate', item['name'], 'found, ignoring...')
        
        # Handle Review data
        elif request.POST.get('import_type') == "Review":
            # Rename columns to match review model
            rename_columns = {"Product Name": "product_name", "Rating": "rating", "Reviewer": "username", "Content": "comment"}
            df.rename(columns = rename_columns, inplace=True)
            reviewers = (df.loc[:,'username']).drop_duplicates() # Get unique reviewers
            # Save all reviewers from the reviews gathered to the seller (CustomUser account='B') model
            for item in reviewers:
                # Need a way to catch verify unique=True error
                try:
                    reviewer = CustomUser(username=item, account='B') # Should not update any existing due to unique constraint
                    reviewer.save()
                except IntegrityError as e:
                    if 'unique constraint' in e.args[0]:
                        # Skip repeats
                        print(e)
            data = df.to_dict('records') # Convert dataframe into dictionaries
            # Loop over list of dictionaries and save each one to the Product model
            for item in data:
                # Find product & seller (CustomUser Account='B') objects
                try:
                    product = Product.objects.get(name=item["product_name"])
                except Product.DoesNotExist:
                    # Skip products not in database
                    print(item["product_name"], "not found!")
                    continue
                try:
                    reviewer = CustomUser.objects.get(username=item["username"], account='B')
                except CustomUser.DoesNotExist:
                    # Skip if cannot find reviewer
                    continue
                review = Review.objects.update_or_create(product_name=product, rating=item["rating"], username=reviewer, comment=item["comment"])
                review[0].save()
        
        return render(request, 'main/import_success.html')
    return render(request, 'main/import_form.html')
        
@login_required(login_url="/login")
@user_passes_test(is_customer, login_url="/home")
def customer_feedback_view(request):
    if request.method == 'POST':
        form = CustomerFeedbackForm(request.POST)
        if form.is_valid():
            feedback = form.save(commit=False)
            feedback.respondent = request.user  # Automatically set the user
            feedback.save()
            # Redirect or display a thank you message
            return render(request, 'main/feedback_thankyou_message.html')

    else:
        form = CustomerFeedbackForm()

    return render(request, 'main/feedback_form.html', {'form': form})


@login_required(login_url="/login")
@user_passes_test(is_seller, login_url="/home")
def seller_feedback_view(request):
    if request.method == 'POST':
        form = SellerFeedbackForm(request.POST)
        if form.is_valid():
            feedback = form.save(commit=False)
            feedback.respondent = request.user  # Automatically set the user
            feedback.save()
            # Redirect or display a thank you message
            return render(request, 'main/feedback_thankyou_message.html')

    else:
        form = SellerFeedbackForm()

    return render(request, 'main/feedback_form.html', {'form': form})


def seller_store(request, username):
    seller = get_object_or_404(CustomUser, username=username)
    products = Product.objects.filter(seller=seller)
    # Calculate the average rating for each product
    for product in products:
        product.average_rating = Review.objects.filter(product_name=product).aggregate(avg_rating=Avg('rating'))['avg_rating']
    context = {
        'seller': seller,
        'products': products,
    }
    return render(request, 'main/seller_store.html', context)


# Common function to update the cart and retrieve cart items
def update_cart(request):
    cart = None
    cartitems = []

    if request.user.is_authenticated:
        cart, created = Cart.objects.get_or_create(user=request.user, completed=False)
        cartitems = cart.cartitems.all()

    context = {"cart": cart, "items": cartitems}
    return context

@login_required(login_url="/login")
@user_passes_test(is_customer, login_url="/home")
def cart(request):
    context = update_cart(request)
    return render(request, 'main/cart.html', context)


@login_required(login_url="/login")
@user_passes_test(is_customer, login_url="/home")
def add_to_cart(request, product_id):
    product = get_object_or_404(Product, product_id=product_id)

    if request.user.is_authenticated:
        cart, created = Cart.objects.get_or_create(user=request.user, completed=False)
        cartitem, created = CartItem.objects.get_or_create(cart=cart, product=product)
        cartitem.quantity += 1
        cartitem.save()
        #messages.success(request, 'You have added an item to your cart successfully')

    # Call the common function to update the cart
    context = update_cart(request)

    return render(request, 'main/cart.html', context)

def remove_from_cart(request, product_id):
    product = get_object_or_404(Product, pk=product_id)
    cart, created = Cart.objects.get_or_create(user=request.user)
    cartitem = CartItem.objects.filter(cart=cart, product=product).first()

    if cartitem:
        cartitem.quantity -= 1

        if cartitem.quantity <= 0:
            cartitem.delete()
        else:
            cartitem.save()

    # Call the common function to update the cart
    context = update_cart(request)

    return render(request, 'main/cart.html', context)

def clear_cart(request):
    cart, created = Cart.objects.get_or_create(user=request.user)
    cart.delete()
    # Call the common function to update the cart
    context = update_cart(request)

    return render(request, 'main/cart.html', context)


def checkout(request):
    if request.method == 'POST':
        payment_form = PaymentForm(request.POST)

        if payment_form.is_valid():
            # Get or create the user's cart
            cart, created = Cart.objects.get_or_create(user=request.user)

            # Create an Order
            order = Order.objects.create(buyer=request.user)

            # Create OrderItems from CartItems
            for cart_item in cart.cartitems.all():
                OrderItem.objects.create(
                    product=cart_item.product,
                    order=order,
                    quantity=cart_item.quantity
                )

            # Clear the cart
            cart.delete()

            # Create a Payment
            payment = Payment.objects.create(order_id=order, completed=True)

            return render(request, 'main/payment_success.html', {'order': order, 'payment': payment})

    else:
        payment_form = PaymentForm()

    context = {'payment_form': payment_form}
    return render(request, 'main/payment.html', context)


def buyer_view_orders(request):
    # Get all orders for the current user
    orders = Order.objects.filter(buyer=request.user).order_by('-order_created_time')
    return render(request, 'main/buyer_view_orders.html', {'orders': orders})

def buyer_view_order_detail(request, order_id):
    # Get the specific order or return a 404 page if not found
    order = get_object_or_404(Order, order_id=order_id, buyer=request.user)

    # Get the order items for the selected order
    order_items = order.orderitems.all()

    return render(request, 'main/buyer_view_order_detail.html', {'order': order, 'order_items': order_items})
